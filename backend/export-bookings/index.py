'''
Business: Export bookings to Excel file for admin download
Args: event with httpMethod GET, optional queryStringParameters (date_from, date_to, room_id)
Returns: Excel file with bookings data
'''

import json
import os
from datetime import datetime
from typing import Dict, Any, List
import psycopg2
from psycopg2.extras import RealDictCursor
from io import BytesIO
import base64

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    openpyxl = None

def get_db_connection():
    database_url = os.environ.get('DATABASE_URL')
    return psycopg2.connect(database_url)

def handler(event: Dict[str, Any], context: Any) -> Dict[str, Any]:
    method: str = event.get('httpMethod', 'GET')
    
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Max-Age': '86400'
            },
            'body': '',
            'isBase64Encoded': False
        }
    
    if not openpyxl:
        return {
            'statusCode': 500,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps({'error': 'openpyxl not installed'}, ensure_ascii=False),
            'isBase64Encoded': False
        }
    
    conn = get_db_connection()
    
    try:
        params = event.get('queryStringParameters') or {}
        date_from = params.get('date_from')
        date_to = params.get('date_to')
        room_id = params.get('room_id')
        
        query = "SELECT * FROM bookings WHERE 1=1"
        query_params = []
        
        if date_from:
            query += " AND created_at >= %s"
            query_params.append(date_from)
        
        if date_to:
            query += " AND created_at <= %s"
            query_params.append(date_to)
        
        if room_id:
            query += " AND room_id = %s"
            query_params.append(room_id)
        
        query += " ORDER BY created_at DESC"
        
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            if query_params:
                cur.execute(query, tuple(query_params))
            else:
                cur.execute(query)
            
            bookings = cur.fetchall()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Бронирования"
        
        header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ["ID", "Номер", "Гость", "Телефон", "Дата заезда", "Дата выезда", "Гостей", "Комментарий", "Статус", "Создано"]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for row_num, booking in enumerate(bookings, 2):
            ws.cell(row=row_num, column=1, value=booking['id'])
            ws.cell(row=row_num, column=2, value=booking['room_name'])
            ws.cell(row=row_num, column=3, value=booking['guest_name'])
            ws.cell(row=row_num, column=4, value=booking['guest_phone'])
            ws.cell(row=row_num, column=5, value=booking['check_in_date'].isoformat() if booking['check_in_date'] else '')
            ws.cell(row=row_num, column=6, value=booking['check_out_date'].isoformat() if booking['check_out_date'] else '')
            ws.cell(row=row_num, column=7, value=booking['guests_count'])
            ws.cell(row=row_num, column=8, value=booking['comment'] or '')
            ws.cell(row=row_num, column=9, value=booking['status'])
            ws.cell(row=row_num, column=10, value=booking['created_at'].isoformat() if booking['created_at'] else '')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        excel_base64 = base64.b64encode(excel_file.read()).decode('utf-8')
        
        filename = f"bookings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': f'attachment; filename="{filename}"',
                'Access-Control-Allow-Origin': '*'
            },
            'body': excel_base64,
            'isBase64Encoded': True
        }
    
    finally:
        conn.close()